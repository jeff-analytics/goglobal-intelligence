from __future__ import annotations

import json

from app import ai_layer, ai_recovery, config, storage


class _FakeResponse:
    def __init__(self, payload): self._payload=payload
    def raise_for_status(self): return None
    def json(self): return self._payload


def _temp_db(tmp_path, monkeypatch):
    monkeypatch.setattr(storage, "DB_PATH", tmp_path / "ai.db")
    storage.init_db()


def test_openai_responses_validation_is_token_free(tmp_path, monkeypatch):
    _temp_db(tmp_path, monkeypatch)
    seen={}
    def fake_get(url, **kwargs):
        seen["url"]=url
        return _FakeResponse({"data":[{"id":"m"}]})
    monkeypatch.setattr(ai_layer.requests,"get",fake_get)
    monkeypatch.setattr(ai_layer.requests,"post",lambda *a,**k: (_ for _ in ()).throw(AssertionError("validation must not generate")))
    out=ai_layer.test_connection({"provider":"X","protocol":"openai_responses","base_url":"https://example.test/v1","api_key":"k","model":"m"})
    assert out["ok"] is True
    assert out["model_generation_used"] is False
    assert seen["url"] == "https://example.test/v1/models"


def test_ai_evidence_store_preserves_provenance(tmp_path, monkeypatch):
    _temp_db(tmp_path, monkeypatch)
    storage.save_ai_evidence({"project_id":7,"market":"DE","evidence_type":"field","field_name":"tariff.rate","value":5.0,"source_name":"Customs","source_url":"https://example.gov/tariff","source_type":"official","evidence_level":"B","retrieval_method":"official-source-ai-extraction","confidence":"high"})
    rows=storage.list_ai_evidence(7,"DE")
    assert len(rows)==1
    assert rows[0]["value"]==5.0
    assert rows[0]["source_url"]=="https://example.gov/tariff"
    assert rows[0]["evidence_level"]=="B"


def test_ai_overlay_does_not_replace_user_tariff_or_tax(tmp_path, monkeypatch):
    _temp_db(tmp_path, monkeypatch)
    storage.save_tariff_override(market="DE",hs_code="850440",rate=3.0,reference_year=2026,note="user")
    storage.save_tax_override(market="DE",rate=19.0,reference_year=2026,note="user")
    snap={"market":"DE","reporter_code":"276","currency":"EUR","hs_code":"850440","origin":{"code":"156","name":"China"},"start_year":2024,"end_year":2025,"trade":{},"tariff":{"rate":None},"fx":{"rate":1.0},"quality":{}}
    project={"id":1,"title":"P","origin":"China","hs_code":"850440","assumptions":{"tax_rate":0.2}}
    saved=[{"field_name":"tariff.rate","value":0.0,"evidence_level":"B","source_name":"AI","source_url":"https://x.gov","retrieval_method":"ai","confidence":"high"},{"field_name":"tax.rate","value":0.0,"evidence_level":"B","source_name":"AI","source_url":"https://x.gov","retrieval_method":"ai","confidence":"high"}]
    out=ai_recovery._apply_evidence(project,snap,saved)
    assert (out.get("tariff") or {}).get("rate") is None
    assert out.get("tax") is None



def test_user_duty_assumption_precedes_recovered_or_source_tariff(tmp_path, monkeypatch):
    _temp_db(tmp_path, monkeypatch)
    from app import main
    project={"id":11,"title":"P","hs_code":"850440","assumptions":{"factory_cost":100,"platform_fee_rate":0.1,"target_margin_rate":0.2,"duty_rate":0.02,"tax_rate":0.0}}
    snap={"market":"DE","tariff":{"rate":9.0},"tax":{"rate":19.0},"fx":{"rate":1.0}}
    result=main._pricing_for_market(project,snap)
    assert result is not None
    assert result.duty_cost == 2.0


def test_ai_supplier_and_local_code_fill_only_missing_snapshot_fields(tmp_path, monkeypatch):
    _temp_db(tmp_path, monkeypatch)
    snap={"market":"DE","reporter_code":"276","currency":"EUR","hs_code":"850440","origin":{"code":"156","name":"China"},"start_year":2024,"end_year":2025,"trade":{},"tariff":{"rate":None},"fx":{"rate":1.0},"quality":{},"suppliers":{},"tariff_official_lookup":{}}
    project={"id":1,"title":"P","origin":"China","hs_code":"850440","assumptions":{}}
    saved=[
        {"field_name":"tariff.local_code","value":"8504408390","evidence_level":"B","source_name":"Customs","source_url":"https://x.gov/tariff","retrieval_method":"web","confidence":"high"},
        {"field_name":"supply.top_suppliers","value":[{"partner_name":"China","iso3":"CHN","trade_value":60,"share":0.6},{"partner_name":"Vietnam","iso3":"VNM","trade_value":40,"share":0.4}],"evidence_level":"B","source_name":"Stats","source_url":"https://x.gov/trade","retrieval_method":"web","confidence":"high"},
        {"field_name":"supply.cr3","value":100,"evidence_level":"B","source_name":"Stats","source_url":"https://x.gov/trade","retrieval_method":"web","confidence":"high"},
    ]
    out=ai_recovery._apply_evidence(project,snap,saved)
    assert out["tariff_official_lookup"]["local_code"] == "8504408390"
    assert out["suppliers"]["supplier_count"] == 2
    assert out["suppliers"]["cr3"] == 1.0


def test_source_backed_ai_marketplace_snapshot_can_be_research_benchmark(tmp_path, monkeypatch):
    _temp_db(tmp_path, monkeypatch)
    from app import main
    storage.save_listing_snapshot({"environment":"ai-web-research","source":"AI evidence recovery","project_id":5,"market_code":"DE","marketplace":"WEB_RESEARCH","query":"charger","verified_market_data":False,"source_backed_market_data":True,"currency":"EUR","evidence_level":"B/C","comparable_set":{"summary":{"median":39.9,"p25":34.9,"p75":44.9},"accepted":[1,2,3]}})
    b=main._research_benchmark({"id":5,"assumptions":{}},"DE")
    assert b is not None
    assert b["median"] == 39.9
    assert b["verified"] is False
    assert b["source_backed"] is True



def test_listing_snapshots_do_not_collide_across_projects_or_markets(tmp_path, monkeypatch):
    _temp_db(tmp_path, monkeypatch)
    storage.save_listing_snapshot({"project_id":1,"market_code":"DE","environment":"csv","marketplace":"UPLOAD","query":"charger","items":[1]})
    storage.save_listing_snapshot({"project_id":2,"market_code":"DE","environment":"csv","marketplace":"UPLOAD","query":"charger","items":[2]})
    storage.save_listing_snapshot({"project_id":1,"market_code":"UK","environment":"csv","marketplace":"UPLOAD","query":"charger","items":[3]})
    rows=storage.list_listing_snapshots()
    assert len(rows)==3
    assert {(r.get("project_id"),r.get("market_code")) for r in rows} == {(1,"DE"),(2,"DE"),(1,"UK")}



def test_explicit_ai_hs_endpoint_returns_candidates_without_overwriting_project(tmp_path, monkeypatch):
    _temp_db(tmp_path, monkeypatch)
    from app import main
    p=storage.create_project({"product_type_id":"generic","title":"Robot vacuum","description":"","origin":"China","hs_code":"","attributes":{},"markets":["DE"],"assumptions":{},"status":"draft"})
    monkeypatch.setattr(main,"recovery_capabilities",lambda: {"configured":True,"native_web_search":True})
    monkeypatch.setattr(main,"recover_hs_candidates",lambda project,query,limit: {"candidates":[{"code":"850811","description":"Vacuum cleaners","relative_confidence":0.8,"source_url":"https://example.gov/hs"}],"count":1})
    out=main.ai_hs_candidates(project_id=p["id"],limit=8)
    assert out["candidates"][0]["code"]=="850811"
    assert storage.get_project(p["id"])["hs_code"]==""


def test_excel_export_contains_ai_evidence_sheet():
    from io import BytesIO
    from openpyxl import load_workbook
    from app.exporter import build_project_workbook
    data=build_project_workbook(project={"id":1,"title":"P"},snapshots=[],decisions=[],listing_snapshots=[],ai_evidence=[{"market":"DE","evidence_type":"field","field_name":"tariff.rate","value":5.0,"source_name":"Customs","source_url":"https://example.gov","evidence_level":"B","retrieval_method":"web","confidence":"high"}])
    wb=load_workbook(BytesIO(data),read_only=True)
    assert "AI Evidence" in wb.sheetnames
    ws=wb["AI Evidence"]
    assert ws.cell(2,3).value == "tariff.rate"


def test_legacy_listing_snapshot_schema_migrates_project_and_market(tmp_path, monkeypatch):
    import json, sqlite3
    db=tmp_path/"legacy.db"
    conn=sqlite3.connect(db)
    conn.execute("CREATE TABLE listing_snapshots (id INTEGER PRIMARY KEY AUTOINCREMENT, environment TEXT NOT NULL, marketplace TEXT NOT NULL, query TEXT NOT NULL, synced_at TEXT NOT NULL, payload_json TEXT NOT NULL, UNIQUE(environment, marketplace, query))")
    payload={"project_id":9,"market_code":"DE","environment":"csv","marketplace":"UPLOAD","query":"charger","items":[]}
    conn.execute("INSERT INTO listing_snapshots(environment,marketplace,query,synced_at,payload_json) VALUES(?,?,?,?,?)",("csv","UPLOAD","charger","2026-08-28T00:00:00+00:00",json.dumps(payload)))
    conn.commit();conn.close()
    monkeypatch.setattr(storage,"DB_PATH",db)
    storage.init_db()
    rows=storage.list_listing_snapshots()
    assert rows[0]["project_id"]==9
    with sqlite3.connect(db) as c:
        cols={r[1] for r in c.execute("PRAGMA table_info(listing_snapshots)")}
    assert {"project_id","market_code"} <= cols


def test_deepseek_compatible_recovery_uses_search_then_structured_responses(monkeypatch):
    calls=[]
    monkeypatch.setattr(ai_recovery, "_config", lambda: {"provider":"DeepSeek","protocol":"openai_compatible","base_url":"https://api.deepseek.com/v1","api_key":"k","model":"deepseek-v4-flash"})
    def fake_post(url, **kwargs):
        body=kwargs.get("json") or {};calls.append((url,body))
        if body.get("tools"):
            return _FakeResponse({"id":"r-search","status":"completed","output":[
                {"type":"web_search_call","status":"completed","action":{"type":"search","query":"Canada tariff"}},
                {"type":"message","status":"completed","role":"assistant","content":[{"type":"output_text","text":"Canada customs lists the applicable rate and tariff line at https://customs.example.gov/tariff"}]}
            ],"usage":{"input_tokens":30,"output_tokens":20,"total_tokens":50}})
        structured={"evidence":[
            {"field":"tariff.rate","value":"0","unit":"percent","source_name":"Canada Customs","source_url":"https://customs.example.gov/tariff","source_type":"official","observed_at":"2026","confidence":"high","excerpt":"rate"},
            {"field":"tariff.local_code","value":"8525890010","unit":"code","source_name":"Canada Customs","source_url":"https://customs.example.gov/tariff","source_type":"official","observed_at":"2026","confidence":"high","excerpt":"code"}
        ],"market_access":[],"marketplace_observations":[],"gaps":[]}
        return _FakeResponse({"id":"r-structure","status":"completed","output":[{"type":"message","status":"completed","role":"assistant","content":[{"type":"output_text","text":json.dumps(structured)}]}],"usage":{"input_tokens":40,"output_tokens":30,"total_tokens":70}})
    monkeypatch.setattr(ai_recovery.requests,"post",fake_post)
    out=ai_recovery._native_web_research({"title":"Camera","origin":"China","hs_code":"852589"},"CA",{"market":"CA","trade":{},"tariff":{},"fx":{},"suppliers":{}},["tariff"])
    assert len(calls)==2
    assert calls[0][0] == "https://api.deepseek.com/responses"
    assert calls[0][1]["tools"] == [{"type":"web_search"}]
    assert calls[0][1]["tool_choice"] == {"type":"web_search"}
    assert calls[1][1]["text"]["format"]["type"] == "json_schema"
    assert "tools" not in calls[1][1]
    assert out["evidence"][0]["field"] == "tariff.rate"
    assert out["_model_calls"]==2
    assert out["_model_usage"]["total_tokens"]==120

def test_recover_market_reports_and_applies_source_backed_result(tmp_path, monkeypatch):
    _temp_db(tmp_path, monkeypatch)
    project=storage.create_project({"product_type_id":"generic","title":"Digital Cameras","description":"","origin":"China","hs_code":"852589","attributes":{"origin_partner_code":"156"},"markets":["CA"],"assumptions":{},"status":"active"})
    snap=storage.save_snapshot({"market":"CA","reporter_code":"124","currency":"CAD","hs_code":"852589","origin":{"code":"156","name":"China"},"start_year":2021,"end_year":2025,"trade":{"latest_total_imports":None,"latest_imports_from_origin":None,"latest_origin_share":None,"history":[]},"suppliers":{},"tariff":{"rate":None},"tariff_official_lookup":{},"fx":{"rate":1.0},"quality":{}})
    monkeypatch.setattr(ai_recovery,"refresh_settings",lambda:None)
    monkeypatch.setattr(ai_recovery,"_supports_native_web",lambda cfg:True)
    monkeypatch.setattr(ai_recovery,"_registered_sources",lambda market:[])
    monkeypatch.setattr(ai_recovery,"_native_web_research",lambda *args,**kwargs:{"evidence":[{"field":"tariff.rate","value":0.0,"source_name":"Canada Customs","source_url":"https://www.cbsa-asfc.gc.ca/tariff","source_type":"official","confidence":"high"},{"field":"tariff.local_code","value":"8525890010","source_name":"Canada Customs","source_url":"https://www.cbsa-asfc.gc.ca/tariff","source_type":"official","confidence":"high"}],"market_access":[],"marketplace_observations":[],"gaps":[],"_model_calls":1,"_model_usage":{"total_tokens":123}})
    out=ai_recovery.recover_market(project,snap,"CA",requested=["tariff"])
    assert out["status"] == "recovered"
    assert out["saved"] == 2
    assert out["applied"] == 2
    assert out["snapshot_updated"] is True
    latest=storage.latest_snapshot("CA","852589",origin_code="156")
    assert latest["tariff"]["rate"] == 0.0
    assert set(latest["ai_recovery"]["applied_fields"]) == {"tariff.rate","tariff.local_code"}


def test_recover_market_returns_visible_no_evidence_state(tmp_path, monkeypatch):
    _temp_db(tmp_path, monkeypatch)
    project=storage.create_project({"product_type_id":"generic","title":"Digital Cameras","description":"","origin":"China","hs_code":"852589","attributes":{"origin_partner_code":"156"},"markets":["CA"],"assumptions":{},"status":"active"})
    snap=storage.save_snapshot({"market":"CA","reporter_code":"124","currency":"CAD","hs_code":"852589","origin":{"code":"156","name":"China"},"start_year":2021,"end_year":2025,"trade":{},"suppliers":{},"tariff":{"rate":None},"tariff_official_lookup":{},"fx":{},"quality":{}})
    monkeypatch.setattr(ai_recovery,"refresh_settings",lambda:None)
    monkeypatch.setattr(ai_recovery,"_supports_native_web",lambda cfg:True)
    monkeypatch.setattr(ai_recovery,"_registered_sources",lambda market:[])
    monkeypatch.setattr(ai_recovery,"_native_web_research",lambda *args,**kwargs:{"evidence":[],"market_access":[],"marketplace_observations":[],"gaps":["tariff"],"_model_calls":1,"_model_usage":{"total_tokens":99}})
    out=ai_recovery.recover_market(project,snap,"CA",requested=["tariff"])
    assert out["status"] == "no_evidence"
    assert out["saved"] == 0
    assert out["snapshot_updated"] is False
    assert out["gaps"] == ["tariff"]


def test_openai_compatible_json_mode_sets_response_format(monkeypatch):
    seen={}
    def fake_post(url, **kwargs):
        seen["url"]=url;seen["json"]=kwargs.get("json")
        return _FakeResponse({"choices":[{"message":{"content":'{"ok":true}'}}]})
    monkeypatch.setattr(ai_layer.requests,"post",fake_post)
    cfg={"provider":"DeepSeek","protocol":"openai_compatible","base_url":"https://api.example.test","api_key":"k","model":"m"}
    _, text, _=ai_layer._post_prompt(cfg,system="Return JSON",user="x",json_mode=True)
    assert seen["json"]["response_format"] == {"type":"json_object"}
    assert ai_layer._parse_json_object(text)["ok"] is True


def test_deepseek_bad_structuring_is_bounded_and_never_loops(monkeypatch):
    cfg={"provider":"DeepSeek","protocol":"openai_compatible","base_url":"https://api.deepseek.com","api_key":"k","model":"deepseek-v4-flash"}
    calls=[]
    def fake_post(url, **kwargs):
        body=kwargs.get("json") or {};calls.append((url,body))
        if body.get("tools"):
            return _FakeResponse({"id":"r1","status":"completed","output":[{"type":"message","status":"completed","role":"assistant","content":[{"type":"output_text","text":"Tariff is 5% according to https://customs.example.gov/tariff"}]}]})
        return _FakeResponse({"id":"r2","status":"completed","output":[{"type":"message","status":"completed","role":"assistant","content":[{"type":"output_text","text":"still not json"}]}]})
    monkeypatch.setattr(ai_recovery.requests,"post",fake_post)
    import pytest
    with pytest.raises(RuntimeError,match="MODEL_INVALID_JSON"):
        ai_recovery._responses_web_research(cfg,system="research",user="x",schema=ai_recovery._base_research_schema())
    # search + Responses JSON Schema + Responses JSON Object + final DeepSeek Chat JSON fallback.
    # Still bounded: no recursive repair loop.
    assert len(calls)==4
    assert calls[0][1].get("tools")==[{"type":"web_search"}]
    assert calls[1][1]["text"]["format"]["type"]=="json_schema"
    assert calls[2][1]["text"]["format"]["type"]=="json_object"
    assert calls[3][0]=="https://api.deepseek.com/chat/completions"
    assert calls[3][1]["response_format"]=={"type":"json_object"}
    assert calls[3][1]["thinking"]=={"type":"disabled"}

def test_deepseek_chat_json_disables_thinking_and_sends_one_call(monkeypatch):
    seen=[]
    def fake_post(url,**kwargs):
        seen.append((url,kwargs.get("json") or {}))
        return _FakeResponse({"choices":[{"message":{"content":'{"ok":true}'}}],"usage":{"prompt_tokens":10,"completion_tokens":5,"total_tokens":15}})
    monkeypatch.setattr(ai_layer.requests,"post",fake_post)
    cfg={"provider":"DeepSeek","protocol":"openai_compatible","base_url":"https://api.deepseek.com/v1","api_key":"k","model":"deepseek-v4-flash"}
    data,text,url=ai_layer._post_prompt(cfg,system="Return JSON",user="x",json_mode=True)
    assert len(seen)==1
    assert url=="https://api.deepseek.com/chat/completions"
    assert seen[0][1]["thinking"]=={"type":"disabled"}
    assert seen[0][1]["response_format"]=={"type":"json_object"}
    assert ai_layer._parse_json_object(text)["ok"] is True


def test_ai_brief_is_one_call_structured_responses_and_reports_usage(monkeypatch):
    calls=[]
    def fake_post(url,**kwargs):
        calls.append((url,kwargs.get("json") or {}))
        content='{"headline":"结论","summary":"摘要","strengths":["优势"],"risks":["风险"],"evidence_gaps":[],"next_actions":["建议"],"decision_language":"有条件推进"}'
        return _FakeResponse({
            "id":"resp_1","status":"completed","model":"deepseek-v4-flash",
            "output":[{"type":"message","status":"completed","role":"assistant","content":[{"type":"output_text","text":content}]}],
            "usage":{"input_tokens":50,"output_tokens":30,"total_tokens":80}
        })
    monkeypatch.setattr(ai_layer,"_config",lambda overrides=None:{"provider":"DeepSeek","protocol":"openai_compatible","base_url":"https://api.deepseek.com","api_key":"k","model":"deepseek-v4-flash"})
    monkeypatch.setattr(ai_layer.requests,"post",fake_post)
    out=ai_layer.generate_evidence_brief(product={"title":"P"},market_contract={"market":"CA"},decision={"status":"CONDITIONAL"},language="zh")
    assert len(calls)==1
    assert calls[0][0]=="https://api.deepseek.com/responses"
    assert calls[0][1]["reasoning"]=={"effort":"none"}
    assert calls[0][1]["text"]["format"]["type"]=="json_schema"
    assert "tools" not in calls[0][1]
    assert out["usage"]["total_tokens"]==80
    assert out["result"]["next_actions"]==["建议"]


def test_deepseek_official_response_shape_recovers_vat(tmp_path, monkeypatch):
    _temp_db(tmp_path, monkeypatch)
    project=storage.create_project({"product_type_id":"generic","title":"Keyboard","description":"","origin":"China","hs_code":"847160","attributes":{"origin_partner_code":"156"},"markets":["GB"],"assumptions":{"tax_rate":0.2},"status":"active"})
    snap=storage.save_snapshot({"market":"GB","reporter_code":"826","currency":"GBP","hs_code":"847160","origin":{"code":"156","name":"China"},"start_year":2024,"end_year":2025,"trade":{"latest_total_imports":10,"latest_origin_share":0.1},"suppliers":{"suppliers":[{"partner_name":"China"}]},"tariff":{"rate":0},"tax":{},"fx":{"rate":0.78},"quality":{}})
    payload={
        "evidence":[{"field":"tax.rate","value":"20","unit":"percent","source_name":"HMRC","source_url":"https://www.gov.uk/vat-rates","source_type":"official","observed_at":"2026","confidence":"high","excerpt":"Standard VAT rate is 20%."}],
        "market_access":[],"marketplace_observations":[],"gaps":[]
    }
    calls=[]
    def fake_post(url,**kwargs):
        body=kwargs.get("json") or {};calls.append(body)
        if body.get("tools"):
            return _FakeResponse({"id":"resp_search","status":"completed","model":"deepseek-v4-flash","output":[
                {"type":"web_search_call","status":"completed","action":{"type":"search","query":"UK VAT rate"}},
                {"type":"message","status":"completed","role":"assistant","content":[{"type":"output_text","text":"HMRC says the standard VAT rate is 20%. https://www.gov.uk/vat-rates"}]}
            ],"usage":{"input_tokens":100,"output_tokens":60,"total_tokens":160}})
        return _FakeResponse({"id":"resp_struct","status":"completed","model":"deepseek-v4-flash","output":[{"type":"message","status":"completed","role":"assistant","content":[{"type":"output_text","text":json.dumps(payload)}]}],"usage":{"input_tokens":80,"output_tokens":40,"total_tokens":120}})
    monkeypatch.setattr(ai_recovery,"_config",lambda:{"provider":"DeepSeek","protocol":"openai_responses","base_url":"https://api.deepseek.com","api_key":"k","model":"deepseek-v4-flash"})
    monkeypatch.setattr(ai_recovery,"_registered_sources",lambda market:[])
    monkeypatch.setattr(ai_recovery.requests,"post",fake_post)
    out=ai_recovery.recover_market(project,snap,"GB",requested=["tax"])
    assert out["status"]=="recovered"
    assert out["model_calls"]==2
    assert out["usage"]["total_tokens"]==280
    latest=storage.latest_snapshot("GB","847160",origin_code="156")
    assert latest["tax"]["rate"]==20.0
    assert latest["tax"]["source"]=="HMRC"
    assert latest["tax"]["reference_year"]=="2026"

