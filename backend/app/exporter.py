from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def _flatten(prefix: str, value: Any, out: dict[str, Any]) -> None:
    if isinstance(value, dict):
        for k, v in value.items():
            _flatten(f"{prefix}.{k}" if prefix else str(k), v, out)
    elif isinstance(value, list):
        out[prefix] = ", ".join(str(x) for x in value)
    else:
        out[prefix] = value


def _style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F2937")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    for col in range(1, ws.max_column + 1):
        max_len = 10
        for row in range(1, min(ws.max_row, 200) + 1):
            value = ws.cell(row=row, column=col).value
            max_len = max(max_len, min(60, len(str(value or "")) + 2))
        ws.column_dimensions[get_column_letter(col)].width = max_len
    ws.freeze_panes = "A2"


def build_project_workbook(*, project: dict[str, Any], snapshots: list[dict[str, Any]], decisions: list[dict[str, Any]], listing_snapshots: list[dict[str, Any]], explorer_rows: list[dict[str, Any]] | None = None, supply_profile: dict[str, Any] | None = None, tariff_matrix: list[dict[str, Any]] | None = None, ai_evidence: list[dict[str, Any]] | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Project"
    flat: dict[str, Any] = {}
    _flatten("", project, flat)
    ws.append(["Field", "Value"])
    for key, value in flat.items():
        ws.append([key, value])
    _style_sheet(ws)

    wm = wb.create_sheet("Market Evidence")
    wm.append(["Market", "Year", "Imports", "Origin imports", "Origin share", "YoY", "CAGR", "Volatility", "CR3", "CR5", "HHI", "Tariff %", "Trade coverage", "Evidence status", "Synced at"])
    for s in snapshots:
        trade = s.get("trade") or {}
        suppliers = s.get("suppliers") or {}
        quality = s.get("quality") or {}
        wm.append([
            s.get("market"), trade.get("latest_year"), trade.get("latest_total_imports"), trade.get("latest_imports_from_origin"), trade.get("latest_origin_share"),
            (trade.get("world_metrics") or {}).get("yoy"), (trade.get("world_metrics") or {}).get("cagr"), trade.get("volatility"), suppliers.get("cr3"), suppliers.get("cr5"), suppliers.get("hhi"),
            (s.get("tariff") or {}).get("rate"), (quality.get("world") or {}).get("coverage_ratio"), quality.get("overall"), s.get("synced_at"),
        ])
    _style_sheet(wm)

    wsup = wb.create_sheet("Suppliers")
    wsup.append(["Market", "Year", "Rank", "Supplier", "ISO3", "Trade value", "Share"])
    for s in snapshots:
        suppliers = s.get("suppliers") or {}
        for rank, row in enumerate(suppliers.get("suppliers") or [], start=1):
            wsup.append([s.get("market"), suppliers.get("year"), rank, row.get("partner_name"), row.get("partner_iso3"), row.get("trade_value"), row.get("share")])
    _style_sheet(wsup)

    we = wb.create_sheet("Opportunity Explorer")
    we.append(["Market", "Latest year", "Imports", "YoY", "3Y CAGR", "Origin share", "Trade coverage", "CR3", "CR5", "HHI", "Tariff", "Evidence completeness", "Pareto frontier", "Quadrant", "Decision status"])
    for r in explorer_rows or []:
        we.append([r.get("market"), r.get("latest_year"), r.get("imports"), r.get("yoy"), r.get("cagr"), r.get("origin_share"), r.get("coverage"), r.get("cr3"), r.get("cr5"), r.get("hhi"), r.get("tariff"), r.get("evidence_ratio"), r.get("pareto_frontier"), r.get("quadrant"), r.get("decision_status")])
    _style_sheet(we)

    wsp = wb.create_sheet("Origin Supply")
    wsp.append(["Section", "Market / destination", "Year / rank", "Trade value", "Share", "CAGR / CR3", "HHI", "Observed", "Source"])
    sp = supply_profile or {}
    sm = sp.get("metrics") or {}
    ds = sp.get("destination_structure") or {}
    if sp:
        wsp.append(["Origin exports", (sp.get("origin") or {}).get("name"), sm.get("latest_year"), sm.get("latest_value"), None, sm.get("cagr"), ds.get("hhi"), True, sp.get("source")])
        for row in sp.get("target_corridors") or []:
            wsp.append(["Target corridor", row.get("market"), row.get("rank"), row.get("trade_value"), row.get("share"), None, None, row.get("observed"), sp.get("source")])
        for row in ds.get("destinations") or []:
            wsp.append(["Destination", row.get("partner_name"), row.get("rank"), row.get("trade_value"), row.get("share"), ds.get("cr3"), ds.get("hhi"), True, sp.get("source")])
    _style_sheet(wsp)

    wtm = wb.create_sheet("Tariff Matrix")
    wtm.append(["Market", "HS6", "Origin code", "Requested year", "Data year", "Reference rate %", "Tariff type", "Fallback", "Status", "Source", "Cached at"])
    for row in tariff_matrix or []:
        wtm.append([row.get("market"), row.get("hs_code"), row.get("origin_code"), row.get("requested_year"), row.get("year"), row.get("rate"), row.get("tariff_type"), row.get("fallback_used"), row.get("status"), row.get("source"), row.get("scanned_at")])
    _style_sheet(wtm)

    wai = wb.create_sheet("AI Evidence")
    wai.append(["Market", "Type", "Field", "Value", "Source", "Source URL", "Level", "Retrieval method", "Confidence", "Observed at", "Retrieved at"])
    for row in ai_evidence or []:
        value = row.get("value")
        if isinstance(value, (dict, list)):
            value = str(value)
        wai.append([row.get("market"), row.get("evidence_type"), row.get("field_name"), value, row.get("source_name"), row.get("source_url"), row.get("evidence_level"), row.get("retrieval_method"), row.get("confidence"), row.get("observed_at"), row.get("retrieved_at")])
    _style_sheet(wai)

    wd = wb.create_sheet("Decision Cases")
    wd.append(["Market", "Status", "Evidence completeness", "Blockers", "Next actions", "Required price", "Benchmark median", "Premium to median"])
    for d in decisions:
        economics = d.get("economics") or {}
        wd.append([
            d.get("market"), d.get("status"), (d.get("evidence_quality") or {}).get("completeness_ratio"),
            " | ".join(d.get("blockers") or []), " | ".join(d.get("next_actions") or []), economics.get("required_price"), economics.get("benchmark_median"), economics.get("premium_to_median"),
        ])
    _style_sheet(wd)

    wl = wb.create_sheet("Listing Snapshots")
    wl.append(["Environment", "Marketplace", "Query", "Source", "Verified", "Total", "Returned", "Synced at"])
    for s in listing_snapshots:
        wl.append([s.get("environment"), s.get("marketplace"), s.get("query"), s.get("source") or "eBay", s.get("verified_market_data"), s.get("total"), s.get("returned"), s.get("synced_at")])
    _style_sheet(wl)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
