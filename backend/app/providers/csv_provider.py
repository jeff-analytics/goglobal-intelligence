from __future__ import annotations

import csv
import io
import re
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .base import ProviderStatus


_FIELD_ALIASES = {
    "item_id": ("item_id", "listing_id", "id", "itemid", "listingid"),
    "title": ("title", "product", "name", "product_title", "item_title", "listing_title"),
    "price": ("price", "listing_price", "amount", "item_price", "sale_price"),
    "currency": ("currency", "currency_code", "ccy"),
    "condition": ("condition", "item_condition"),
    "category_id": ("category_id", "category", "categoryid"),
    "brand": ("brand", "brand_name"),
    "shipping_cost": ("shipping_cost", "shipping", "delivery_cost", "postage"),
    "seller": ("seller", "seller_name", "merchant"),
    "url": ("url", "item_url", "listing_url", "web_url"),
    "item_location": ("item_location", "location", "seller_location"),
    "observed_at": ("observed_at", "snapshot_date", "date", "captured_at"),
}


def _normalized_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def _get(row: dict[str, Any], field: str) -> Any:
    for key in _FIELD_ALIASES[field]:
        if key in row and row.get(key) not in (None, ""):
            return row.get(key)
    return None


def _number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    # Accept common marketplace exports such as "$1,299.00", "EUR 49.90".
    text = text.replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def _infer_currency(price_value: Any, currency_value: Any) -> str | None:
    if currency_value not in (None, ""):
        return str(currency_value).strip().upper()
    text = str(price_value or "")
    if "€" in text:
        return "EUR"
    if "£" in text:
        return "GBP"
    if "A$" in text or "AUD" in text.upper():
        return "AUD"
    if "C$" in text or "CAD" in text.upper():
        return "CAD"
    if "US$" in text.upper() or "USD" in text.upper() or "$" in text:
        return "USD"
    return None


def _rows_from_csv(payload: bytes) -> Iterable[tuple[int, dict[str, Any]]]:
    text = payload.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    for i, row in enumerate(reader, start=2):
        yield i, {_normalized_header(k): v for k, v in row.items()}


def _rows_from_xlsx(payload: bytes) -> Iterable[tuple[int, dict[str, Any]]]:
    wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    try:
        ws = wb.active
        iterator = ws.iter_rows(values_only=True)
        header_row = next(iterator, None)
        if not header_row:
            return
        headers = [_normalized_header(x) for x in header_row]
        for i, values in enumerate(iterator, start=2):
            row = {headers[idx]: value for idx, value in enumerate(values) if idx < len(headers) and headers[idx]}
            yield i, row
    finally:
        wb.close()


class CsvProvider:
    key = "csv"
    name = "Uploaded marketplace observations"

    def status(self) -> ProviderStatus:
        return ProviderStatus(
            key=self.key,
            name=self.name,
            configured=True,
            environment="user-upload",
            supports_taxonomy=False,
            supports_search=False,
            supports_market_benchmark=True,
            note="",
        )

    def parse_bytes(self, payload: bytes, *, filename: str = "upload.csv") -> list[dict[str, Any]]:
        suffix = Path(filename).suffix.lower()
        if suffix == ".csv":
            raw_rows = _rows_from_csv(payload)
        elif suffix in {".xlsx", ".xlsm"}:
            raw_rows = _rows_from_xlsx(payload)
        else:
            raise ValueError("Supported formats: CSV, XLSX")

        rows: list[dict[str, Any]] = []
        for i, lower in raw_rows:
            if not any(value not in (None, "") for value in lower.values()):
                continue
            raw_price = _get(lower, "price")
            title = str(_get(lower, "title") or "").strip()
            rows.append({
                "item_id": str(_get(lower, "item_id") or f"row-{i}"),
                "title": title,
                "price": _number(raw_price),
                "currency": _infer_currency(raw_price, _get(lower, "currency")),
                "condition": str(_get(lower, "condition") or ""),
                "category_id": str(_get(lower, "category_id") or "") or None,
                "brand": str(_get(lower, "brand") or "") or None,
                "shipping_cost": _number(_get(lower, "shipping_cost")),
                "seller": str(_get(lower, "seller") or "") or None,
                "url": str(_get(lower, "url") or "") or None,
                "item_location": str(_get(lower, "item_location") or "") or None,
                "observed_at": str(_get(lower, "observed_at") or "") or None,
                "source_row": i,
            })
        if not rows:
            raise ValueError("No marketplace observation rows found")
        return rows
