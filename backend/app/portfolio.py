from __future__ import annotations

import csv
import io
import uuid
from typing import Any

from openpyxl import load_workbook


FIELD_ALIASES = {
    "title": {"title", "product", "product_name", "name", "sku_name"},
    "sku": {"sku", "sku_id", "item_sku"},
    "origin": {"origin", "country_of_origin", "origin_country"},
    "hs_code": {"hs", "hs_code", "customs_code", "tariff_code"},
    "markets": {"markets", "target_markets", "market"},
    "factory_cost": {"factory_cost", "cost", "unit_cost", "cogs"},
    "packaging_cost": {"packaging_cost", "packaging"},
    "freight_cost": {"freight_cost", "freight", "shipping_cost"},
    "fulfillment_cost": {"fulfillment_cost", "fulfillment"},
    "platform_fee_rate": {"platform_fee_rate", "platform_fee", "platform_fee_pct"},
    "target_margin_rate": {"target_margin_rate", "target_margin", "margin", "margin_pct"},
    "base_currency": {"base_currency", "currency"},
}


def _norm_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_").replace("-", "_")


def _mapped(row: dict[str, Any]) -> dict[str, Any]:
    source = {_norm_header(k): v for k, v in row.items()}
    out: dict[str, Any] = {}
    for dest, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            if alias in source and source[alias] not in (None, ""):
                out[dest] = source[alias]
                break
    return out


def _float(value: Any) -> float | None:
    try:
        if value in (None, ""):
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def _rate(value: Any) -> float | None:
    val = _float(value)
    if val is None:
        return None
    # Human spreadsheets commonly use 15 for 15%. Accept either convention.
    return val / 100 if val > 1 else val


def _market_list(value: Any) -> list[str]:
    if value in (None, ""):
        return []
    text = str(value)
    parts = [x.strip().upper() for sep in [";", ",", "|", "/"] for x in []]
    import re
    return [x.strip().upper() for x in re.split(r"[;,|/]", text) if x.strip()]


def normalize_portfolio_row(raw: dict[str, Any], *, row_number: int) -> dict[str, Any]:
    row = _mapped(raw)
    title = str(row.get("title") or "").strip()
    errors = []
    if len(title) < 2:
        errors.append("Product name is required")
    assumptions = {
        "base_currency": str(row.get("base_currency") or "").strip().upper() or None,
        "factory_cost": _float(row.get("factory_cost")),
        "packaging_cost": _float(row.get("packaging_cost")),
        "freight_cost": _float(row.get("freight_cost")),
        "fulfillment_cost": _float(row.get("fulfillment_cost")),
        "platform_fee_rate": _rate(row.get("platform_fee_rate")),
        "target_margin_rate": _rate(row.get("target_margin_rate")),
    }
    assumptions = {k: v for k, v in assumptions.items() if v is not None}
    return {
        "row_number": row_number,
        "sku": str(row.get("sku") or "").strip() or None,
        "title": title,
        "origin": str(row.get("origin") or "").strip(),
        "hs_code": "".join(ch for ch in str(row.get("hs_code") or "") if ch.isalnum()),
        "markets": _market_list(row.get("markets")),
        "assumptions": assumptions,
        "errors": errors,
    }


def parse_portfolio_bytes(payload: bytes, *, filename: str) -> list[dict[str, Any]]:
    lower = filename.lower()
    rows: list[dict[str, Any]] = []
    if lower.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        ws = wb.active
        values = ws.iter_rows(values_only=True)
        try:
            headers = [_norm_header(x) for x in next(values)]
        except StopIteration:
            return []
        for idx, vals in enumerate(values, start=2):
            raw = {headers[i]: vals[i] if i < len(vals) else None for i in range(len(headers))}
            if not any(v not in (None, "") for v in raw.values()):
                continue
            rows.append(normalize_portfolio_row(raw, row_number=idx))
    else:
        text = payload.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for idx, raw in enumerate(reader, start=2):
            if not any(v not in (None, "") for v in raw.values()):
                continue
            rows.append(normalize_portfolio_row(raw, row_number=idx))
    return rows


def portfolio_batch_id() -> str:
    return uuid.uuid4().hex[:12]
